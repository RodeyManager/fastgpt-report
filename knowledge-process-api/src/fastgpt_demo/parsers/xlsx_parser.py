"""XLSX parser — reads all sheets with openpyxl, outputs CSV + Markdown tables."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

from ._types import ParseResult

CUSTOM_SPLIT_SIGN = "-----CUSTOM_SPLIT_SIGN-----"


def _sheet_to_csv_rows(ws) -> list[list[str]]:
    """Convert a worksheet to a list of string rows (CSV-style)."""
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        str_row = [str(cell) if cell is not None else "" for cell in row]
        rows.append(str_row)
    return rows


def _rows_to_csv(rows: list[list[str]]) -> str:
    """Write rows as CSV text."""
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return buf.getvalue()


def _rows_to_markdown_table(rows: list[list[str]]) -> str:
    """Convert rows to a Markdown table string."""
    if not rows:
        return ""

    header = rows[0]
    col_count = len(header)

    def _sanitize(cell: str) -> str:
        return cell.replace("\n", "\\n")

    lines: list[str] = [
        f"| {' | '.join(_sanitize(c) for c in header)} |",
        f"| {' | '.join(['---'] * col_count)} |",
    ]

    for row in rows[1:]:
        padded = (row + [""] * col_count)[:col_count]
        lines.append(f"| {' | '.join(_sanitize(c) for c in padded)} |")

    return "\n".join(lines)


def parse(buffer: bytes, **kwargs) -> ParseResult:
    """Parse an XLSX buffer — all sheets, CSV raw text + Markdown tables."""

    wb = load_workbook(filename=BytesIO(buffer), read_only=True, data_only=True)

    sheet_names: list[str] = wb.sheetnames

    csv_parts: list[str] = []
    md_parts: list[str] = []

    for name in sheet_names:
        ws = wb[name]
        rows = _sheet_to_csv_rows(ws)

        csv_text = _rows_to_csv(rows)
        csv_parts.append(csv_text)

        md_table = _rows_to_markdown_table(rows)
        md_parts.append(md_table)

    wb.close()

    raw_text = "\n".join(csv_parts)
    format_text = CUSTOM_SPLIT_SIGN.join(md_parts)

    return ParseResult(
        raw_text=raw_text,
        format_text=format_text,
        html_preview=f"<pre>{raw_text}</pre>",
        image_list=[],
        sheet_names=sheet_names,
    )
